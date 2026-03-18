#!/usr/bin/env python3
"""Vorderland Brain – Dokument-Vorbereitung fuer die Ingestion-Pipeline.

Konvertiert und splittet grosse Dateien fuer das Embedding.
Unterstuetzte Formate: PDF, XLSX, DOCX, TXT, MD, CSV

Qualitaetscheck am Ende prueft:
- Vollstaendigkeit (alle Blaetter/Seiten uebernommen)
- Inhaltsdichte (keine ueberfluessigen Leerzeilen)
- Anomalien (Timestamps, Encoding-Probleme)
"""

import sys
import os
import re
import shutil
from pathlib import Path

# Schwellwerte (in Bytes)
TEXT_MAX_SIZE = 50_000      # 50 KB fuer Textdateien
BINARY_MAX_SIZE = 100_000   # 100 KB fuer PDF, XLSX, DOCX
SPLIT_TARGET_SIZE = 40_000  # ~40 KB pro Teil
OVERLAP_LINES = 10          # Anzahl Zeilen Ueberlappung zwischen Teilen

# Qualitaetscheck-Schwellwerte
MAX_EMPTY_LINE_RATIO = 0.30  # Max 30% Leerzeilen
MIN_CONTENT_CHARS = 100      # Mindestens 100 Zeichen Inhalt

# Textbasierte Formate die direkt gelesen werden koennen
TEXT_FORMATS = {".txt", ".md", ".csv"}
# Binaere Formate die konvertiert werden muessen
BINARY_FORMATS = {".pdf", ".xlsx", ".xls", ".docx"}


def find_inbox(start_path=None):
    """Findet den Inbox-Ordner im Vorderland Brain Projekt."""
    if start_path:
        p = Path(start_path)
        while p != p.parent:
            inbox = p / "docker" / "shared" / "inbox"
            if inbox.parent.exists():
                inbox.mkdir(exist_ok=True)
                return inbox
            p = p.parent

    # Fallback: Bekannte Projektpfade (Windows OneDrive + macOS)
    fallbacks = [
        Path.home() / "OneDrive - Region Vorderland-Feldkirch" / "Claude" / "Dev" / "vorderland-brain" / "docker" / "shared" / "inbox",
        Path.home() / "Library" / "CloudStorage" / "OneDrive-RegionVorderland-Feldkirch" / "Claude" / "Dev" / "vorderland-brain" / "docker" / "shared" / "inbox",
    ]
    for known in fallbacks:
        if known.parent.exists():
            known.mkdir(exist_ok=True)
            return known

    return None


def extract_text_from_pdf(file_path):
    """Extrahiert Text aus PDF mit pdfplumber oder PyPDF2."""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
        return "\n\n".join(text_parts)
    except ImportError:
        pass

    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(file_path)
        text_parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
        return "\n\n".join(text_parts)
    except ImportError:
        pass

    print("FEHLER: Weder pdfplumber noch PyPDF2 installiert.")
    print("Installiere mit: pip install pdfplumber")
    sys.exit(1)


def get_xlsx_sheet_names(file_path):
    """Gibt die Blattnamen einer XLSX-Datei zurueck."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        return wb.sheetnames
    except ImportError:
        return []


def extract_text_from_xlsx(file_path):
    """Extrahiert Text aus Excel-Dateien mit Bereinigung."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"=== Blatt: {sheet_name} ===\n")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells).strip()
                if line and line != "|":
                    text_parts.append(line)
        return "\n".join(text_parts)
    except ImportError:
        print("FEHLER: openpyxl nicht installiert.")
        print("Installiere mit: pip install openpyxl")
        sys.exit(1)


def extract_text_from_docx(file_path):
    """Extrahiert Text aus Word-Dokumenten."""
    try:
        import docx
        doc = docx.Document(file_path)
        text_parts = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(text_parts)
    except ImportError:
        print("FEHLER: python-docx nicht installiert.")
        print("Installiere mit: pip install python-docx")
        sys.exit(1)


def read_text_file(file_path):
    """Liest eine Textdatei."""
    encodings = ["utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            return Path(file_path).read_text(encoding=enc)
        except (UnicodeDecodeError, UnicodeError):
            continue
    print(f"FEHLER: Konnte {file_path} mit keinem Encoding lesen.")
    sys.exit(1)


# --- Textbereinigung ---

def is_empty_pipe_line(line):
    """Erkennt leere Pipe-Zeilen wie '|  |  |  |  |'."""
    stripped = line.strip()
    if not stripped:
        return True
    # Nur Pipes und Whitespace
    cleaned = stripped.replace("|", "").replace(" ", "")
    return len(cleaned) == 0


def clean_text(text):
    """Bereinigt extrahierten Text fuer bessere Embedding-Qualitaet.

    - Entfernt leere Pipe-Zeilen (aus Excel)
    - Bereinigt Timestamps (00:00:00 entfernen)
    - Entfernt aufeinanderfolgende Leerzeilen (max 1)
    - Entfernt Leerzeilen am Ende von Blaettern
    """
    lines = text.split("\n")
    cleaned = []
    prev_empty = False

    for line in lines:
        # Timestamp bereinigen: "2023-08-09 00:00:00" -> "2023-08-09"
        line = re.sub(r'(\d{4}-\d{2}-\d{2})\s+00:00:00', r'\1', line)

        # Leere Pipe-Zeilen ueberspringen
        if is_empty_pipe_line(line):
            if not prev_empty:
                # Maximal eine Leerzeile behalten (als Absatztrenner)
                cleaned.append("")
            prev_empty = True
            continue

        prev_empty = False
        cleaned.append(line)

    # Leerzeilen vor "=== Blatt:" auf genau eine reduzieren
    result = []
    for i, line in enumerate(cleaned):
        if line.startswith("=== Blatt:") and result:
            # Leerzeilen vor Blatt-Marker auf eine reduzieren
            while result and result[-1] == "":
                result.pop()
            result.append("")
        result.append(line)

    # Trailing-Leerzeilen entfernen
    while result and result[-1] == "":
        result.pop()

    return "\n".join(result)


# --- Qualitaetscheck ---

def quality_check(file_path, text, targets, suffix):
    """Prueft die Qualitaet der Konvertierung.

    Returns:
        (ok, fehler, warnungen, verbesserungen)
        ok: True wenn bestanden
        fehler: Liste kritischer Fehler (fuehren zum Abbruch)
        warnungen: Liste von Warnungen (kein Abbruch)
        verbesserungen: Liste von Verbesserungsvorschlaegen
    """
    fehler = []
    warnungen = []
    verbesserungen = []

    # 1. Grundcheck: Text nicht leer
    if not text or len(text.strip()) < MIN_CONTENT_CHARS:
        fehler.append(f"Extrahierter Text ist zu kurz ({len(text.strip())} Zeichen, Minimum: {MIN_CONTENT_CHARS})")
        return False, fehler, warnungen, verbesserungen

    lines = text.split("\n")
    total_lines = len(lines)
    content_lines = [l for l in lines if l.strip()]
    empty_lines = total_lines - len(content_lines)

    # 2. Leerzeilen-Verhaeltnis
    if total_lines > 0:
        empty_ratio = empty_lines / total_lines
        if empty_ratio > MAX_EMPTY_LINE_RATIO:
            fehler.append(
                f"Zu viele Leerzeilen: {empty_lines}/{total_lines} ({empty_ratio:.0%}). "
                f"Maximum: {MAX_EMPTY_LINE_RATIO:.0%}. "
                f"Textbereinigung hat nicht korrekt funktioniert."
            )
            verbesserungen.append("Leere Pipe-Zeilen aus Excel wurden nicht vollstaendig entfernt")

    # 3. XLSX-spezifisch: Blaetter pruefen
    if suffix in (".xlsx", ".xls"):
        source_sheets = get_xlsx_sheet_names(file_path)
        # Alle Texte der Zieldateien zusammenlesen
        combined_text = ""
        for t in targets:
            combined_text += Path(t).read_text(encoding="utf-8")

        found_sheets = re.findall(r'=== Blatt: (.+?) ===', combined_text)

        missing = set(source_sheets) - set(found_sheets)
        extra = set(found_sheets) - set(source_sheets)

        if missing:
            fehler.append(f"Fehlende Blaetter in TXT: {', '.join(sorted(missing))}")
            verbesserungen.append("Alle Blaetter muessen konvertiert werden, auch leere")

        if extra:
            warnungen.append(f"Unbekannte Blaetter in TXT: {', '.join(sorted(extra))}")

        # Inhalt pro Blatt pruefen
        for sheet in source_sheets:
            pattern = f"=== Blatt: {re.escape(sheet)} ==="
            if pattern.replace("\\", "") not in combined_text:
                continue
            # Suche Text zwischen diesem Blatt-Marker und dem naechsten
            sheet_pattern = re.escape(f"=== Blatt: {sheet} ===")
            sheet_match = re.search(
                sheet_pattern + r'\n(.*?)(?==== Blatt:|$)',
                combined_text,
                re.DOTALL
            )
            if sheet_match:
                sheet_content = sheet_match.group(1).strip()
                if len(sheet_content) < 10:
                    warnungen.append(f"Blatt '{sheet}' hat sehr wenig Inhalt ({len(sheet_content)} Zeichen)")

    # 4. Encoding-Probleme erkennen
    encoding_issues = re.findall(r'[ï¿½ð]|\\x[0-9a-f]{2}', text)
    if encoding_issues:
        warnungen.append(f"Moeglicherweise Encoding-Probleme gefunden ({len(encoding_issues)} Stellen)")
        verbesserungen.append("Encoding der Quelldatei pruefen (UTF-8 vs. Latin-1)")

    # 5. Uebrig gebliebene leere Pipe-Zeilen zaehlen
    remaining_pipes = sum(1 for l in lines if re.match(r'^\s*(\|\s*)+\s*$', l))
    if remaining_pipes > 5:
        warnungen.append(f"Noch {remaining_pipes} leere Pipe-Zeilen im Text vorhanden")
        verbesserungen.append("Filter fuer leere Pipe-Zeilen erweitern")

    # 6. Dateigroesse der Zieldateien pruefen
    for t in targets:
        t_size = Path(t).stat().st_size
        if t_size > SPLIT_TARGET_SIZE * 1.5:
            warnungen.append(f"{Path(t).name} ist {t_size/1024:.0f} KB (Ziel: ~{SPLIT_TARGET_SIZE/1024:.0f} KB)")

    ok = len(fehler) == 0
    return ok, fehler, warnungen, verbesserungen


def repair_text(text, fehler, warnungen):
    """Versucht den Text automatisch zu reparieren basierend auf den Fehlern/Warnungen.

    Returns:
        (reparierter_text, aenderungen) – Liste der durchgefuehrten Reparaturen
    """
    aenderungen = []
    lines = text.split("\n")

    # Reparatur 1: Zu viele Leerzeilen → aggressiver filtern
    for item in fehler + warnungen:
        if "Leerzeilen" in item or "leere Pipe-Zeilen" in item:
            new_lines = []
            prev_empty = False
            for line in lines:
                stripped = line.strip()
                # Aggressiver: Auch Zeilen mit nur Nummern und Pipes entfernen
                cleaned = stripped.replace("|", "").replace(" ", "")
                if not cleaned or (cleaned.isdigit() and len(cleaned) <= 3 and "|" in stripped):
                    if not prev_empty:
                        new_lines.append("")
                    prev_empty = True
                    continue
                prev_empty = False
                new_lines.append(line)
            removed = len(lines) - len(new_lines)
            if removed > 0:
                lines = new_lines
                aenderungen.append(f"Aggressive Leerzeilen-Bereinigung: {removed} weitere Zeilen entfernt")

    # Reparatur 2: Timestamps nochmal pruefen
    repaired_lines = []
    ts_fixed = 0
    for line in lines:
        new_line = re.sub(r'(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}:\d{2}', r'\1', line)
        if new_line != line:
            ts_fixed += 1
        repaired_lines.append(new_line)
    if ts_fixed > 0:
        lines = repaired_lines
        aenderungen.append(f"Timestamps bereinigt: {ts_fixed} Stellen")

    # Reparatur 3: Encoding-Artefakte entfernen
    repaired_lines = []
    enc_fixed = 0
    for line in lines:
        new_line = re.sub(r'[ï¿½]', '', line)
        if new_line != line:
            enc_fixed += 1
        repaired_lines.append(new_line)
    if enc_fixed > 0:
        lines = repaired_lines
        aenderungen.append(f"Encoding-Artefakte entfernt: {enc_fixed} Stellen")

    # Trailing-Leerzeilen entfernen
    while lines and lines[-1] == "":
        lines.pop()

    return "\n".join(lines), aenderungen


def print_quality_report(ok, fehler, warnungen, verbesserungen, runde=1):
    """Gibt den Qualitaetsbericht aus."""
    print()
    print("=" * 50)
    label = f"QUALITAETSCHECK (Runde {runde})" if runde > 1 else "QUALITAETSCHECK"
    print(label)
    print("=" * 50)

    if fehler:
        print()
        print("FEHLER:")
        for f in fehler:
            print(f"  X {f}")

    if warnungen:
        print()
        print("WARNUNGEN:")
        for w in warnungen:
            print(f"  ! {w}")

    if verbesserungen:
        print()
        print("VERBESSERUNGSVORSCHLAEGE:")
        for v in verbesserungen:
            print(f"  > {v}")

    if ok and not warnungen:
        print()
        print("  Alle Pruefungen bestanden.")

    print()
    if ok:
        print("Ergebnis: BESTANDEN")
    else:
        print("Ergebnis: NICHT BESTANDEN – Reparatur wird versucht...")

    print("=" * 50)


# --- Hilfsfunktionen ---

def is_section_boundary(line):
    """Erkennt ob eine Zeile eine Abschnittsgrenze ist."""
    stripped = line.strip()
    if not stripped:
        return False

    if re.match(r'^#{1,4}\s', stripped):
        return True
    if re.match(r'^[=\-*]{3,}\s*$', stripped):
        return True
    if re.match(r'^xx\d+[\.\d]*:', stripped):
        return True
    if stripped.startswith("=== Blatt:"):
        return True
    if re.match(r'^Seite\s+\d+\s+von\s+\d+', stripped):
        return True
    if re.match(r'^\d+\.\d*\s+[A-ZÄÖÜ]', stripped):
        return True

    return False


def split_text(text, target_size=SPLIT_TARGET_SIZE, overlap=OVERLAP_LINES):
    """Teilt Text in Abschnitte von ~target_size Bytes."""
    lines = text.split("\n")
    parts = []
    current_lines = []
    current_size = 0
    last_boundary_idx = 0
    last_boundary_size = 0

    for i, line in enumerate(lines):
        line_size = len(line.encode("utf-8")) + 1

        if is_section_boundary(line) and current_size > target_size * 0.3:
            last_boundary_idx = len(current_lines)
            last_boundary_size = current_size

        if current_size + line_size > target_size and current_lines:
            if last_boundary_idx > 0 and last_boundary_size > target_size * 0.3:
                split_at = last_boundary_idx
                parts.append("\n".join(current_lines[:split_at]))
                overlap_start = max(0, split_at - overlap)
                current_lines = current_lines[overlap_start:]
                current_size = sum(len(l.encode("utf-8")) + 1 for l in current_lines)
            else:
                parts.append("\n".join(current_lines))
                overlap_start = max(0, len(current_lines) - overlap)
                overlap_lines = current_lines[overlap_start:]
                current_lines = list(overlap_lines)
                current_size = sum(len(l.encode("utf-8")) + 1 for l in current_lines)

            last_boundary_idx = 0
            last_boundary_size = 0

        current_lines.append(line)
        current_size += line_size

    if current_lines:
        parts.append("\n".join(current_lines))

    return parts


# --- Hauptfunktion ---

def prepare_document(file_path, inbox_path):
    """Hauptfunktion: Dokument vorbereiten, bereinigen, pruefen und in Inbox legen."""
    file_path = Path(file_path).resolve()

    if not file_path.exists():
        print(f"FEHLER: Datei nicht gefunden: {file_path}")
        sys.exit(1)

    suffix = file_path.suffix.lower()
    base_name = file_path.stem
    file_size = file_path.stat().st_size

    print(f"Datei: {file_path.name}")
    print(f"Groesse: {file_size / 1024:.0f} KB")
    print(f"Format: {suffix}")
    print()

    # Schwellwert bestimmen
    max_size = TEXT_MAX_SIZE if suffix in TEXT_FORMATS else BINARY_MAX_SIZE

    # Kleine Textdateien: direkt kopieren (kein Qualitaetscheck noetig)
    if suffix in TEXT_FORMATS and file_size <= max_size:
        target = inbox_path / file_path.name
        shutil.copy2(file_path, target)
        print(f"Datei ist klein genug ({file_size / 1024:.0f} KB <= {max_size / 1024:.0f} KB).")
        print(f"Kopiert nach: {target}")
        return [target]

    # PDFs: direkt kopieren (n8n-Pipeline kann PDFs selbst verarbeiten)
    if suffix == ".pdf" and file_size <= max_size:
        target = inbox_path / file_path.name
        shutil.copy2(file_path, target)
        print(f"PDF ist klein genug ({file_size / 1024:.0f} KB <= {max_size / 1024:.0f} KB).")
        print(f"Kopiert nach: {target}")
        return [target]

    # XLSX/DOCX: IMMER zu Text konvertieren (Embeddings API akzeptiert nur Text)
    if suffix in (".xlsx", ".xls", ".docx"):
        print(f"Binaerformat {suffix} wird zu Text konvertiert (Embeddings API braucht Text)...")
        print()

    # Grosse PDFs/Textdateien: Text extrahieren und splitten
    if suffix == ".pdf":
        print(f"PDF ist zu gross ({file_size / 1024:.0f} KB > {max_size / 1024:.0f} KB). Wird aufgeteilt...")
        print()

    if suffix == ".pdf":
        text = extract_text_from_pdf(file_path)
    elif suffix in (".xlsx", ".xls"):
        text = extract_text_from_xlsx(file_path)
    elif suffix == ".docx":
        text = extract_text_from_docx(file_path)
    elif suffix in TEXT_FORMATS:
        print(f"Textdatei ist zu gross ({file_size / 1024:.0f} KB > {max_size / 1024:.0f} KB). Wird aufgeteilt...")
        print()
        text = read_text_file(file_path)
    else:
        print(f"FEHLER: Format {suffix} wird nicht unterstuetzt.")
        print(f"Unterstuetzte Formate: {', '.join(sorted(TEXT_FORMATS | BINARY_FORMATS))}")
        sys.exit(1)

    if not text.strip():
        print("WARNUNG: Kein Text extrahiert. Die Datei ist moeglicherweise leer oder nur Bilder.")
        sys.exit(1)

    raw_size = len(text.encode("utf-8"))
    raw_lines = len(text.split("\n"))
    print(f"Extrahierter Text (roh): {raw_size / 1024:.0f} KB, {raw_lines} Zeilen")

    # Bereinigung
    text = clean_text(text)
    clean_size = len(text.encode("utf-8"))
    clean_lines = len(text.split("\n"))
    removed = raw_lines - clean_lines
    print(f"Nach Bereinigung: {clean_size / 1024:.0f} KB, {clean_lines} Zeilen ({removed} Zeilen entfernt)")
    print()

    # Aufteilen
    parts = split_text(text)

    if len(parts) == 1:
        target = inbox_path / f"{base_name}.txt"
        target.write_text(parts[0], encoding="utf-8")
        print(f"Text in eine Datei geschrieben: {target}")
        targets = [target]
    else:
        targets = []
        for i, part in enumerate(parts, 1):
            target = inbox_path / f"{base_name}-teil{i}.txt"
            target.write_text(part, encoding="utf-8")
            part_size = len(part.encode("utf-8"))
            targets.append(target)
            print(f"  Teil {i}/{len(parts)}: {target.name} ({part_size / 1024:.0f} KB)")
        print()
        print(f"{len(parts)} Teile in Inbox geschrieben.")

    # Qualitaetscheck mit Reparatur-Schleife (max 3 Versuche)
    MAX_REPAIR_ROUNDS = 3
    for runde in range(1, MAX_REPAIR_ROUNDS + 1):
        ok, fehler, warnungen, verbesserungen = quality_check(
            file_path, text, targets, suffix
        )
        print_quality_report(ok, fehler, warnungen, verbesserungen, runde=runde)

        if ok:
            break

        if runde >= MAX_REPAIR_ROUNDS:
            print()
            print(f"ABBRUCH: Nach {MAX_REPAIR_ROUNDS} Reparaturversuchen nicht bestanden.")
            print("Die Dateien wurden aus der Inbox entfernt.")
            print("Bitte die Quelldatei manuell pruefen.")
            for t in targets:
                try:
                    Path(t).unlink()
                except OSError:
                    pass
            sys.exit(1)

        # Reparatur versuchen
        print()
        print(f"Reparatur (Versuch {runde})...")
        text, aenderungen = repair_text(text, fehler, warnungen)

        if not aenderungen:
            print("  Keine automatische Reparatur moeglich.")
            print("  Die Dateien wurden aus der Inbox entfernt.")
            for t in targets:
                try:
                    Path(t).unlink()
                except OSError:
                    pass
            sys.exit(1)

        for a in aenderungen:
            print(f"  + {a}")

        # Alte Dateien loeschen und neu schreiben
        for t in targets:
            try:
                Path(t).unlink()
            except OSError:
                pass

        clean_size = len(text.encode("utf-8"))
        clean_lines = len(text.split("\n"))
        print(f"  Text nach Reparatur: {clean_size / 1024:.0f} KB, {clean_lines} Zeilen")

        parts = split_text(text)
        if len(parts) == 1:
            target = inbox_path / f"{base_name}.txt"
            target.write_text(parts[0], encoding="utf-8")
            targets = [target]
        else:
            targets = []
            for i, part in enumerate(parts, 1):
                target = inbox_path / f"{base_name}-teil{i}.txt"
                target.write_text(part, encoding="utf-8")
                targets.append(target)

    return targets


def main():
    if len(sys.argv) < 2:
        print("Verwendung: python prepare.py <datei-pfad> [inbox-pfad]")
        print()
        print("Beispiele:")
        print("  python prepare.py /pfad/zur/datei.pdf")
        print("  python prepare.py datei.xlsx /pfad/zum/inbox/")
        sys.exit(1)

    file_path = sys.argv[1]

    if len(sys.argv) >= 3:
        inbox_path = Path(sys.argv[2])
    else:
        inbox_path = find_inbox(os.getcwd())

    if not inbox_path:
        print("FEHLER: Inbox-Ordner nicht gefunden.")
        print("Gib den Pfad als zweites Argument an.")
        sys.exit(1)

    inbox_path.mkdir(parents=True, exist_ok=True)
    print(f"Inbox: {inbox_path}")
    print("=" * 50)

    targets = prepare_document(file_path, inbox_path)

    print()
    print(f"Fertig! {len(targets)} Datei(en) in der Inbox.")
    print("Starte die Ingestion ueber das Dashboard oder den Webhook.")


if __name__ == "__main__":
    main()
